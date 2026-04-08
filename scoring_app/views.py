from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Avg
from django.http import HttpResponse
from .models import Event, ScoringItem, Team, Commissioner, ScoreRecord, TeamRankRecord
from decimal import Decimal
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def home(request):
    return render(request, 'scoring_app/home.html')

@login_required
def setup_event(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        passcode = request.POST.get('passcode')
        is_anonymous = request.POST.get('is_anonymous_results') == 'on'
        
        # Check if passcode already exists
        if Event.objects.filter(passcode=passcode).exists():
            messages.error(request, f'代碼「{passcode}」已被使用，請更換一個代碼。')
            # Collect items to pass back
            item_names = request.POST.getlist('item_name')
            item_weights = request.POST.getlist('item_weight')
            items_data = zip(item_names, item_weights)
            
            return render(request, 'scoring_app/setup_event.html', {
                'title': title,
                'description': description,
                'passcode': passcode,
                'is_anonymous': is_anonymous,
                'commissioner_names': request.POST.getlist('commissioner_name'),
                'team_names': request.POST.get('team_names'),
                'items_data': items_data
            })

        event = Event.objects.create(
            title=title,
            description=description,
            passcode=passcode,
            is_anonymous_results=is_anonymous
        )
        
        # Items (1-6)
        item_names = request.POST.getlist('item_name')
        item_weights = request.POST.getlist('item_weight')
        for name, weight in zip(item_names, item_weights):
            if name and weight:
                ScoringItem.objects.create(event=event, name=name, weight=int(weight))
                
        # Teams
        team_names = request.POST.get('team_names').splitlines()
        for name in team_names:
            if name.strip():
                Team.objects.create(event=event, name=name.strip())
                
        # Generate commissioners from input list
        comm_names = request.POST.getlist('commissioner_name')
        if not comm_names:
            comm_names = ["委員 1"]
            
        for i, name in enumerate(comm_names, 1):
            c_name = name.strip()
            if not c_name:
                c_name = f"委員 {i}"
            Commissioner.objects.create(event=event, name=c_name)
            
        return redirect('event_dashboard', event_id=event.id)
        
    return render(request, 'scoring_app/setup_event.html', {'commissioner_names': ['']})

@login_required
def event_list(request):
    events = Event.objects.all().order_by('-created_at')
    return render(request, 'scoring_app/event_list.html', {'events': events})

@login_required
def event_dashboard(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    commissioners = event.commissioners.all()
    
    comm_status = []
    completed_count = 0
    for comm in commissioners:
        is_completed = ScoreRecord.objects.filter(commissioner=comm).exists()
        if is_completed:
            completed_count += 1
        comm_status.append({
            'comm': comm,
            'is_completed': is_completed
        })
        
    return render(request, 'scoring_app/event_dashboard.html', {
        'event': event,
        'commissioners': comm_status,
        'completed_count': completed_count,
        'total_count': commissioners.count()
    })

def enter_passcode(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    if request.method == 'POST':
        passcode = request.POST.get('passcode')
        if passcode == event.passcode:
            # For simplicity, we just pick the first available commissioner 
            # or allow them to select their name.
            # User requirement: "Direct link to scoring table" or "passcode".
            # If they enter from a generic link, they might need to select who they are.
            return render(request, 'scoring_app/select_commissioner.html', {
                'event': event,
                'commissioners': event.commissioners.all()
            })
        else:
            messages.error(request, '代碼錯誤')
    return render(request, 'scoring_app/enter_passcode.html', {'event': event})

def scoring_matrix(request, token):
    commissioner = get_object_or_404(Commissioner, token=token)
    
    if ScoreRecord.objects.filter(commissioner=commissioner).exists():
        messages.success(request, '您已經成功完成評分，感謝您的參與！以下為目前的統計結果。')
        return redirect(f"/results/{commissioner.event.id}/?token={token}")

    event = commissioner.event
    teams = event.teams.all()
    items = event.items.all()
    
    if request.method == 'POST':
        # Clear existing scores for this commissioner
        ScoreRecord.objects.filter(commissioner=commissioner).delete()
        TeamRankRecord.objects.filter(commissioner=commissioner).delete()
        
        for team in teams:
            # Save scores for each item
            for item in items:
                score_val = request.POST.get(f'score_{team.id}_{item.id}')
                if score_val:
                    ScoreRecord.objects.create(
                        commissioner=commissioner,
                        team=team,
                        item=item,
                        score=score_val
                    )
            # Save ranking
            rank_val = request.POST.get(f'rank_{team.id}')
            if rank_val:
                TeamRankRecord.objects.create(
                    commissioner=commissioner,
                    team=team,
                    rank=int(rank_val)
                )
        
        return redirect(f"/results/{event.id}/?token={commissioner.token}")

    return render(request, 'scoring_app/scoring_matrix.html', {
        'commissioner': commissioner,
        'event': event,
        'teams': teams,
        'items': items
    })

def event_results(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Permission Check: allow if admin OR if provided a valid commissioner token
    is_admin = request.user.is_authenticated
    token = request.GET.get('token')
    if not is_admin:
        if token:
            if not Commissioner.objects.filter(token=token, event=event).exists():
                return redirect('login')
        else:
            return redirect('login')
    teams = event.teams.all()
    items = event.items.all()
    commissioners = list(event.commissioners.all())
    
    for i, comm in enumerate(commissioners, 1):
        comm.display_name = f"委員 {i}" if event.is_anonymous_results else comm.name
    
    results = []
    for team in teams:
        rank_sum = TeamRankRecord.objects.filter(team=team).aggregate(Sum('rank'))['rank__sum'] or 0
        total_weighted_scores = []
        comm_details = {}
        
        for comm in commissioners:
            comm_scores = ScoreRecord.objects.filter(commissioner=comm, team=team)
            weighted_sum = 0
            if comm_scores.exists():
                for s in comm_scores:
                    weighted_sum += float(s.score) * (float(s.item.weight) / 100.0)
                total_weighted_scores.append(weighted_sum)
            comm_details[comm.display_name] = weighted_sum
            
        avg_score = sum(total_weighted_scores) / len(total_weighted_scores) if total_weighted_scores else 0
        
        results.append({
            'team': team,
            'rank_sum': rank_sum,
            'avg_score': avg_score,
            'comm_details': comm_details
        })
    
    sorted_results = sorted(results, key=lambda x: (x['rank_sum'], -x['avg_score']))
    
    chart_data = {
        "labels": [res['team'].name for res in sorted_results],
        "overall_ranks": [res['rank_sum'] for res in sorted_results],
        "overall_scores": [float(res['avg_score']) for res in sorted_results],
        "comm_scores": {}
    }
    for comm in commissioners:
        chart_data["comm_scores"][comm.display_name] = []
        for res in sorted_results:
            chart_data["comm_scores"][comm.display_name].append(res['comm_details'][comm.display_name])
    
    return render(request, 'scoring_app/results.html', {
        'event': event,
        'results': sorted_results,
        'commissioners': commissioners,
        'chart_data_json': json.dumps(chart_data),
        'is_admin': is_admin
    })

@login_required
def export_results_excel(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    teams = event.teams.all()
    commissioners = list(event.commissioners.all())
    items = list(event.items.all())
    
    results = []
    for team in teams:
        rank_sum = TeamRankRecord.objects.filter(team=team).aggregate(Sum('rank'))['rank__sum'] or 0
        total_weighted_scores = []
        comm_scores_list = []
        comm_ranks_list = []
        
        for comm in commissioners:
            # Score processing
            comm_scores = ScoreRecord.objects.filter(commissioner=comm, team=team)
            weighted_sum = 0
            if comm_scores.exists():
                for s in comm_scores:
                    weighted_sum += float(s.score) * (float(s.item.weight) / 100.0)
                total_weighted_scores.append(weighted_sum)
            comm_scores_list.append(weighted_sum)
            
            # Rank processing
            comm_rank = TeamRankRecord.objects.filter(commissioner=comm, team=team).first()
            comm_ranks_list.append(comm_rank.rank if comm_rank else "-")
            
        avg_score = sum(total_weighted_scores) / len(total_weighted_scores) if total_weighted_scores else 0
        results.append({
            'team': team,
            'team_name': team.name,
            'rank_sum': rank_sum,
            'avg_score': float(avg_score),
            'comm_scores': comm_scores_list,
            'comm_ranks': comm_ranks_list
        })
    
    sorted_results = sorted(results, key=lambda x: (x['rank_sum'], -x['avg_score']))
    
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Overall Summary ---
    ws_summary = wb.active
    ws_summary.title = "總表 (總分與統整序位)"
    
    headers = ["最終排名", "組別名稱", "序位總和", "平均加權總分"]
    for i, comm in enumerate(commissioners, 1):
        d_name = f"委員 {i}" if event.is_anonymous_results else comm.name
        headers.append(f"{d_name} 加權得分")
        headers.append(f"{d_name} 給予序位")
    ws_summary.append(headers)
    
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    for i, res in enumerate(sorted_results, 1):
        row = [i, res['team_name'], res['rank_sum'], res['avg_score']]
        # Interleave scores and ranks for each commissioner
        for c_idx in range(len(commissioners)):
            row.append(res['comm_scores'][c_idx])
            row.append(res['comm_ranks'][c_idx])
        ws_summary.append(row)
        
    for col in ws_summary.columns:
        ws_summary.column_dimensions[col[0].column_letter].width = 18

    # --- Sheet 2..N: Commissioner Details ---
    for i, comm in enumerate(commissioners, 1):
        d_name = f"委員 {i}" if event.is_anonymous_results else comm.name
        ws_comm = wb.create_sheet(title=f"{d_name} 評分明細")
        comm_headers = ["組別名稱"]
        for item in items:
            comm_headers.append(f"{item.name} ({item.weight}%)")
        comm_headers.extend(["總計加權得分", "給予序位"])
        ws_comm.append(comm_headers)
        
        for cell in ws_comm[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            
        for res in sorted_results:
            team = res['team']
            row = [team.name]
            weighted_total = 0
            
            # Fetch each item score
            for item in items:
                record = ScoreRecord.objects.filter(commissioner=comm, team=team, item=item).first()
                if record:
                    val = float(record.score)
                    row.append(val)
                    weighted_total += val * (float(item.weight) / 100.0)
                else:
                    row.append("-")
            
            # Totals & Ranks
            row.append(weighted_total)
            rank_record = TeamRankRecord.objects.filter(commissioner=comm, team=team).first()
            row.append(rank_record.rank if rank_record else "-")
            
            ws_comm.append(row)
            
        for col in ws_comm.columns:
            ws_comm.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # Set proper filename encoding for chinese characters if needed, or stick to ascii
    response['Content-Disposition'] = f'attachment; filename=results_{event.id}.xlsx'
    wb.save(response)
    return response

@login_required
def delete_event(request, event_id):
    if request.method == 'POST':
        event = get_object_or_404(Event, id=event_id)
        title = event.title
        event.delete()
        messages.success(request, f'事件 "{title}" 已成功刪除。')
    return redirect('event_list')

